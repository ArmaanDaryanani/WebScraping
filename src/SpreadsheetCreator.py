from openpyxl import Workbook, load_workbook


class SpreadsheetCreator:

    def __init__(self, filename):
        self.filename = filename
        self.current_row = 1
        # Ensure the file exists; if not, create a new workbook
        try:
            wb = load_workbook(self.filename)
        except FileNotFoundError:
            wb = Workbook()
            wb.save(self.filename)

    def appendDataToSheet(self, sheetname, data):
        wb = load_workbook(self.filename)

        if sheetname not in wb.sheetnames:
            ws = wb.create_sheet(sheetname)
        else:
            ws = wb[sheetname]

        for col_num, value in enumerate(data, 1):
            ws.cell(row = self.current_row, column = col_num, value = value)

        wb.save(self.filename)

    def nextRow(self):
        self.current_row +=1
        print(f"Row: {self.current_row}")

    def resetSheetData(self, sheetname):
        wb = load_workbook(self.filename)

        if sheetname not in wb.sheetnames:
            print(f"'{sheetname}' not found in the workbook {self.filename}. Creating it...")
            ws = wb.create_sheet(sheetname)
        else:
            ws = wb[sheetname]
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

        wb.save(self.filename)

    def deleteSheet(self, sheetname):
        wb = load_workbook(self.filename)

        if sheetname in wb.sheetnames:
            ws = wb[sheetname]
            wb.remove(ws)
            wb.save(self.filename)
            print(f"Sheet '{sheetname}' deleted successfully")
        else:
            print(f"Sheet '{sheetname}' not found")

    def createNewSheet(self, sheetname, init_data):
        wb = load_workbook(self.filename)

        if sheetname not in wb.sheetnames:
            ws = wb.create_sheet(sheetname)
            ws.append(init_data)
            print(f"Sheet '{sheetname}' created.")
        else:
            print(f"'{sheetname}' already exists in the workbook {self.filename}.")

        wb.save(self.filename)

    def autosizeColumns(self, sheetname):
        wb = load_workbook(self.filename)
        ws = wb[sheetname]

        for col in ws.columns:
            max_length = 0
            column = [cell for cell in col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) #to add extra space
            ws.column_dimensions[cell.column_letter].width = adjusted_width

        wb.save(self.filename)
